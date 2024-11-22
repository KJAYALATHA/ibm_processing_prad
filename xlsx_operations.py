import datetime
import os
import pathlib

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook


def get_fo_number_from_jms_file_read(file_path):
    """
    method to read the jms excel file workbook and get only visible sheet Frame Order No
    :param file_path: jms excel file path
    :return: string Frame Order Number
    """
    value = ""
    xl = None
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(file_path, sheet_name=sheet.title, dtype=str)
                n_df = df.to_numpy()
                nd_df = n_df[2][2:3]
                value = nd_df[0]  # <-- based on the JMS input sheet template, if it changes this needs to change
                break
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get FO number from the JMS input sheet, error was {}".format(e))
    return value


# print(get_fo_number_from_jms_file_read(
#     "/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx"))


def get_service_codes_numbers_from_jms_file_read(file_path):
    """
    method to read the jms excel file workbook and get only visible sheet service order numbers
    :param file_path: jms excel file path
    :return: array of service order numbers
    """
    new_list = []
    xl = None
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(file_path, sheet_name=sheet.title, dtype=str)
                np_df = df.to_numpy()[7]  # <-- column in df has all the service codes
                n_df = np_df[5::4]  # <-- based on the JMS input sheet format, if that changes, this needs to be checked
                for item in n_df:
                    str_value = str(item)
                    new_list.append(str_value)
        # new_list.sort(reverse=True)  # <-- delete this line after the calculator test
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get service codes from the JMS input file, error was :{}".format(e))
    finally:
        xl.close()
    return new_list


# print(get_service_codes_numbers_from_jms_file_read(
#     "/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx"))


def check_color_code_and_apply_formula(file_path):
    """
    method to check the cell color code and apply the formula
    :param file_path: excel file absolute path
    :return: None
    """
    xl = pd.ExcelFile(file_path)
    sheets = xl.book.worksheets
    for sheet in sheets:
        # check if the worksheet is visible, we are currently considering only visible worksheets
        if sheet.sheet_state != 'hidden':
            for cell in sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column):
                color_in_hex = cell[10].fill.bgColor.index  # < ---need to fix the cell index to iterate the range
                print('HEX =', color_in_hex)
                print('RGB =', tuple(int(color_in_hex[i:i + 2], 16) for i in (0, 2, 4)))
                break
            break


# check_color_code_and_apply_formula(r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx")


def get_start_and_end_date_from_jms(file_path):
    """
    method to get the start and end date from the jms input sheet
    :param file_path:
    :return: array of start, end dates and month as string
    """
    new_list = []
    month = ""
    xl = None
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(file_path, sheet_name=sheet.title, dtype=str)
                n_df = df.to_numpy()
                arr = n_df[4][2:4]
                for item in arr:
                    new_t = str(item)[:10]
                    month_dt = datetime.datetime.strptime(new_t, "%Y-%m-%d")
                    month = month_dt.strftime("%B")
                    # converting the date format to format compatible with Reliance Back Up Sheet entry format
                    t = datetime.datetime.strptime(new_t, "%Y-%m-%d").strftime("%d.%m.%Y")
                    new_list.append(t)
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get start and end dates from the JMS input file, error was : {}".format(e))
    return new_list, month


# print(get_start_and_end_date_from_jms(
#     r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx"))


def get_jms_number(file_path):
    """
    method to get the JMS number from the JMS input file
    :param file_path:
    :return: JMS Number as string
    """
    value = ""
    xl = None
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(file_path, sheet_name=sheet.title, dtype=str)
                n_df = df.to_numpy()
                value = n_df[3][2:3]
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get the JMS number from the JMS input file, error was :{}".format(e))
    return value


# print(get_jms_number(r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx"))


def set_jms_number(file_path, jms_draft_id):
    """
    method to set the JMS number on to JMS input excel file
    :param file_path:
    :param jms_draft_id:
    :return: None
    """
    sheet_name = None
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                sheet_name = sheet.title
                break
        xl.close()
        # load excel file
        workbook = load_workbook(filename=file_path)
        # open workbook
        sheet = workbook[sheet_name]
        sheet["C5"] = jms_draft_id  # <-- based on current template row is 5
        workbook.save(filename=file_path)
        workbook.close()
        return True
    except Exception as e:
        raise ValueError("Failed to update the JMS value on to the JMS input file, error was :{}".format(e))


# 1108627 - old number
# set_jms_number(r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx",
#                "1108627")

def copy_temp_jms(source_jms_file_path):
    """
    method to copy the jms input sheet to new temporary excel file
    :param source_jms_file_path:
    :return: Saved file path and sheet name
    """
    sheet_name = ""
    try:
        fold_path = pathlib.Path(source_jms_file_path).parent
        new_file_path = os.path.join(fold_path, "NewTextBook.xlsx")
        xl = pd.ExcelFile(source_jms_file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                sheet_name = sheet.title
                break
        xl.close()
        # create an empty workbook with sheet name
        wb = xlsxwriter.Workbook(new_file_path)
        wb.add_worksheet(sheet_name)
        wb.close()
        # loading source file
        source = load_workbook(filename=source_jms_file_path, keep_vba=True, data_only=True)
        wbs = source[sheet_name]
        # loading destination file
        dest = load_workbook(filename=new_file_path)
        wbt = dest[sheet_name]
        # copying all the rows & column values to new file
        for row in wbs:
            for cell in row:
                wbt[cell.coordinate].value = cell.value
        # saving the destination file with copied data
        dest.save(new_file_path)
        dest.close()
    except Exception as e:
        raise ValueError("Failed to copy source jms to temp jsm file, error was :{}".format(e))
    return new_file_path, sheet_name


# copy_temp_jms("/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/JMS Oct 2021 Karnataka.xlsx")


def get_man_month_for_service_code(temp_jms_file_path):
    """
    method to get the man month for service codes
    :param temp_jms_file_path:
    :return: list of service codes in the same order of jms input excel file
    """
    value = []
    try:
        xl = pd.ExcelFile(temp_jms_file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(temp_jms_file_path, sheet_name=sheet.title, skiprows=10,
                                   dtype=str)  # < --skiprows=10
                n_df = df.astype(str).to_numpy()
                np_df = n_df[len(n_df) - 2][6::4]  # <-- currently considering this based on JMS input sheet format
                for item in np_df:
                    str_value = "{:.3f}".format(float(item))  # <-- considering up to 4 decimal places
                    value.append(str_value)
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get the JMS number from the JMS input file, error was :{}".format(e))
    return value


# print(get_man_month_for_service_code(
#     r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/TextBook.xlsx"))


def get_amount(temp_jms_file_path):
    """
    method to get the amount total by service code in the same order of jms input file
    :param temp_jms_file_path:
    :return: list of amount values in the same order of jms input file
    """
    value = []
    try:
        xl = pd.ExcelFile(temp_jms_file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(temp_jms_file_path, sheet_name=sheet.title, skiprows=10)  # < --skiprows=10
                n_df = df.astype(str).to_numpy()
                np_df = n_df[len(n_df) - 2][8::4]  # <-- currently considering this based on JMS input sheet format
                for item in np_df:
                    str_value = "{:.3f}".format(float(item))  # <-- considering up to 4 decimal places
                    value.append(str_value)
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get the JMS number from the JMS input file, error was :{}".format(e))
    return value


# print(get_amount(r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/TextBook.xlsx"))


def get_grand_total(temp_jms_file_path):
    """
    mthod to get the grand total value from the jms input excel file
    :param temp_jms_file_path:
    :return: string of the grand total amount
    """
    value = ""
    try:
        xl = pd.ExcelFile(temp_jms_file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(temp_jms_file_path, sheet_name=sheet.title, skiprows=10)  # < --skiprows=10
                n_df = df.astype(str).to_numpy()
                np_df = n_df[len(n_df) - 1]  # <-- currently considering this based on JMS input sheet format
                new_val = np_df[2:3]
                value = "{:.3f}".format(float(new_val))  # <-- considering up to 4 decimal places
        xl.close()
    except Exception as e:
        raise ValueError("Failed to get the JMS number from the JMS input file, error was :{}".format(e))
    return value


# print(get_grand_total(r"/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/TextBook.xlsx"))

def value_loc(file_path, search_pattern, new_value):
    try:
        # loading destination file
        dest = load_workbook(filename=file_path)
        wbt = dest.active
        # copying all the rows & column values to new file
        for row in wbt:
            for cell in row:
                if wbt[cell.coordinate].value == search_pattern:
                    wbt[cell.coordinate].value = new_value
                    break
        # saving the destination file with copied data
        dest.save(file_path)
        dest.close()
    except Exception as e:
        raise ValueError("Failed to locate {} in the JMS input file, error was :{}".format(search_pattern, e))


def check_value_and_update_formula(file_path, search_pattern, new_value):
    global sheet
    try:
        xl = pd.ExcelFile(file_path)
        sheets = xl.book.worksheets
        for sheet in sheets:
            # check if the worksheet is visible, we are currently considering only visible worksheets
            if sheet.sheet_state != 'hidden':
                df = pd.read_excel(file_path, sheet_name=sheet.title)
                n_df = df.astype(str).to_numpy()
                if search_pattern in n_df:
                    value_loc(file_path, search_pattern, new_value)
                    break
        xl.close()
    except Exception as e:
        raise ValueError("Failed to locate: {} from the JMS input file, error was :{}".format(search_pattern, e))

# check_value_and_update_formula("/Users/pradeeprajusibbal/PycharmProjects/reliance-bot/resources/NewTextBook.xlsx",
#                                "${grand_total}", "=C1+D1+E1+F1")
